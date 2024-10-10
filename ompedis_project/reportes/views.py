from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from django.http import HttpResponse, JsonResponse, FileResponse
from django.db.models import Count, Func
from django.db import models
from datetime import date, datetime, timedelta
from io import BytesIO
from base64 import b64decode
import json
import logging  # Para depuración y registro
from django.views.generic.edit import UpdateView, DeleteView

from .models import SesionTerapia
from .forms import SesionTerapiaForm, ReporteGeneracionForm

# Para la generación de PDFs
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# Para la generación y manipulación de archivos Excel
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image  # Para insertar imágenes en el archivo Excel

# Otros
from collections import Counter
from PIL import Image as PILImage  # Para el manejo de imágenes
from django.core.files.base import ContentFile
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin

# Vistas que utilicen estos imports se pueden definir a continuación


@login_required
def menu_reportes_view(request):
    return render(request, 'reportes/menu_reportes.html')

# Definición de la clase Week para usar EXTRACT en PostgreSQL
class Week(Func):
    function = 'EXTRACT'
    template = "%(function)s('week' FROM %(expressions)s)"
    output_field = models.FloatField()  # Especificar el tipo de campo de salida

#vista para registrar una nueva sesión de terapia
@login_required
def registrar_sesion_view(request):
    if request.method == 'POST':
        form = SesionTerapiaForm(request.POST)
        if form.is_valid():
            sesion = form.save(commit=False)
            sesion.genero = sesion.paciente.genero  # Asignar el género desde el paciente seleccionado
            sesion.save()
            # Agregar un mensaje de éxito
            messages.success(request, 'Se registró la sesión de terapia con éxito.')
            return redirect('registrar_sesion')
        else:
            messages.error(request, 'Hubo un error al registrar la sesión de terapia. Por favor, verifique los datos ingresados.')
    else:
        form = SesionTerapiaForm()

    context = {
        'form': form,
    }
    return render(request, 'reportes/registrar_sesion.html', context)
    
#vista para ver las estadísticas de las sesiones de terapia    
@login_required
def ver_estadisticas_view(request):
    # Filtrar todas las sesiones de terapia con filtros de fecha
    sesiones = SesionTerapia.objects.all()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio:
        sesiones = sesiones.filter(fecha_sesion__gte=fecha_inicio)
    if fecha_fin:
        sesiones = sesiones.filter(fecha_sesion__lte=fecha_fin)

    sesiones = sesiones.order_by('fecha_sesion')    
    # Calcular rangos de edad para los pacientes
    rangos_edad = {
        '0_18': sum(1 for sesion in sesiones if sesion.paciente.calcular_edad() <= 18),
        '19_35': sum(1 for sesion in sesiones if 19 <= sesion.paciente.calcular_edad() <= 35),
        '36_60': sum(1 for sesion in sesiones if 36 <= sesion.paciente.calcular_edad() <= 60),
        '60_plus': sum(1 for sesion in sesiones if sesion.paciente.calcular_edad() > 60),
    }

    # Contar el total de pacientes femeninos y masculinos
    total_femeninos = sesiones.filter(genero='Femenino').count()
    total_masculinos = sesiones.filter(genero='Masculino').count()

    # Obtener el número de terapias por semana
    terapias_por_semana = sesiones.annotate(
        week=Week('fecha_sesion')
    ).values('week').annotate(total=Count('id')).order_by('week')

    # Pasar los datos al contexto
    context = {
        'total_femeninos': total_femeninos,
        'total_masculinos': total_masculinos,
        'rangos_edad': rangos_edad,
        'terapias_por_semana': terapias_por_semana,
        'sesiones': sesiones,  # Añadir las sesiones al contexto para la tabla
        'fecha_inicio': fecha_inicio,  # Añadir fecha_inicio al contexto
        'fecha_fin': fecha_fin,  # Añadir fecha_fin al contexto
    }

    return render(request, 'reportes/ver_estadisticas.html', context)

@login_required
def ver_listado_sesiones_view(request):
    sesiones = SesionTerapia.objects.all()

    context = {
        'sesiones': sesiones,
    }
    return render(request, 'reportes/listado_sesiones.html', context)

@login_required
def generar_reporte_view(request):
    if request.method == 'POST':
        form = ReporteGeneracionForm(request.POST)
        if form.is_valid():
            # Filtrar sesiones de terapia con los parámetros del formulario
            sesiones = SesionTerapia.objects.all()

            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            paciente = form.cleaned_data['paciente']

            if fecha_inicio:
                sesiones = sesiones.filter(fecha_sesion__gte=fecha_inicio)
            if fecha_fin:
                sesiones = sesiones.filter(fecha_sesion__lte=fecha_fin)
            if paciente:
                sesiones = sesiones.filter(paciente=paciente)

            sesiones = sesiones.order_by('fecha_sesion')

            # Crear buffer
            buffer = BytesIO()

            # Crear el PDF
            pdf = SimpleDocTemplate(buffer, pagesize=A4)

            # Estilos de texto y elementos
            styles = getSampleStyleSheet()
            title_style = styles["Title"]
            normal_style = styles["BodyText"]

            # Contenido del PDF
            content = []

            # Encabezado
            content.append(Paragraph("Reporte de Actividades", title_style))
            content.append(Spacer(1, 12))

            # Mostrar nombre completo del paciente si está seleccionado
            if paciente:
                nombre_completo = f"{paciente.nombre} {paciente.apellido or ''} {paciente.nombre} {paciente.apellido or ''}".strip()
                content.append(Paragraph(f"Paciente: {nombre_completo}", normal_style))
                content.append(Spacer(1, 12))

            # Añadir información general
            content.append(Paragraph("Este informe resume las actividades realizadas durante las sesiones de terapia.", normal_style))
            content.append(Spacer(1, 12))

            # Crear tabla de datos dinámicamente a partir de las sesiones filtradas
            data = [['Fecha', 'Diagnóstico', 'Área', 'Género']]  # Encabezados de la tabla

            for sesion in sesiones:
                data.append([
                    sesion.fecha_sesion.strftime('%Y-%m-%d'),
                    sesion.diagnostico,
                    sesion.area,
                    'Masculino' if sesion.genero == 'Masculino' else 'Femenino'
                ])

            # Estilo para la tabla
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ])

            table = Table(data)
            table.setStyle(table_style)

            content.append(table)
            content.append(Spacer(1, 12))

            # Firma final o pie de página
            content.append(Spacer(1, 48))
            content.append(Paragraph("Este reporte fue generado automáticamente por el sistema OMPEDIS.", normal_style))

            # Construir PDF
            pdf.build(content)

            # Enviar PDF como respuesta
            buffer.seek(0)
            return FileResponse(buffer, as_attachment=True, filename='reporte_formal.pdf')
    else:
        form = ReporteGeneracionForm()

    return render(request, 'reportes/generar_reporte.html', {'form': form})

def exportar_excel_view(request): 


    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

        # Convertir fechas a objetos de fecha si están presentes
    try:
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio = None
        fecha_fin = None


    # Filtrar sesiones de terapia con los parámetros de fecha
    sesiones = SesionTerapia.objects.all()
    if fecha_inicio:
        sesiones = sesiones.filter(fecha_sesion__gte=fecha_inicio)
    if fecha_fin:
        sesiones = sesiones.filter(fecha_sesion__lte=fecha_fin)

    sesiones = sesiones.order_by('fecha_sesion')

    # Crear un archivo Excel en memoria
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Estadísticas OMPEDIS'

    # Definir estilos comunes
    header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    header_font = Font(bold=True, size=12)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Crear el encabezado principal
    sheet.merge_cells('A1:H1')
    sheet['A1'] = "UNIVERSIDAD MARIANO GALVEZ DE GUATEMALA"
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A1'].alignment = Alignment(horizontal='center')

    sheet.merge_cells('A2:H2')
    sheet['A2'] = "ESTADÍSTICA OMPEDIS OSTUNCALCO"
    sheet['A2'].font = Font(bold=True, size=14)
    sheet['A2'].alignment = Alignment(horizontal='center')

    sheet.merge_cells('A3:H3')
    sheet['A3'] = f"PERIODO: {fecha_inicio} - {fecha_fin}"
    sheet['A3'].font = Font(bold=True, size=12)
    sheet['A3'].alignment = Alignment(horizontal='center')

    # Crear encabezados de la tabla principal
    headers = ['No.', 'Nombre', 'Diagnóstico', 'Sexo', 'Edad', 'Área', 'Fecha de Inicio']
    sheet.append(headers)
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_style

    # Obtener datos reales de la base de datos
    pacientes = [
        {
            'nombre': f'{sesion.paciente.nombre} {sesion.paciente.apellido}',
            'diagnostico': sesion.diagnostico,
            'sexo': sesion.get_genero_display(),
            'edad': sesion.calcular_edad(),
            'area': sesion.area,
            'fecha_ingreso': sesion.fecha_sesion.strftime('%d/%m/%Y')
        }
        for sesion in sesiones
    ]

    # Agregar datos reales a la tabla principal
    for idx, paciente in enumerate(pacientes, start=1):
        sheet.append([idx, paciente['nombre'], paciente['diagnostico'], paciente['sexo'], paciente['edad'], paciente['area'], paciente['fecha_ingreso']])

    # Aplicar estilo a las celdas de la tabla principal
    for row in sheet.iter_rows(min_row=5, max_row=4 + len(pacientes), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_style

    # Calcular totales de género
    total_femeninos = sesiones.filter(genero='Femenino').count()
    total_masculinos = sesiones.filter(genero='Masculino').count()

    # Crear sección de Resumen de Pacientes por Género
    sheet['J1'] = "Total pacientes femeninos y masculinos"
    sheet['J1'].font = Font(bold=True, size=12)
    sheet['J1'].alignment = Alignment(horizontal='center')

    sheet['J2'] = "Femenino"
    sheet['K2'] = total_femeninos
    sheet['J3'] = "Masculino"
    sheet['K3'] = total_masculinos
    sheet['J4'] = "Total"
    sheet['K4'] = total_femeninos + total_masculinos

    # Estilo para la tabla de género
    for col in range(10, 12):
        for row in range(1, 5):
            cell = sheet.cell(row=row, column=col)
            cell.border = border_style

    # Crear gráfico de barras para género
    chart_genero = BarChart()
    data_genero = Reference(sheet, min_col=11, min_row=2, max_row=3)
    categories_genero = Reference(sheet, min_col=10, min_row=2, max_row=3)
    chart_genero.add_data(data_genero, titles_from_data=False)
    chart_genero.set_categories(categories_genero)
    chart_genero.title = "Total pacientes femeninos y masculinos"
    chart_genero.x_axis.title = "Sexo"
    chart_genero.y_axis.title = "Total"
    sheet.add_chart(chart_genero, "M1")

    # Calcular rango de edades
    rangos_edad = {
        '0 a 10 años': sesiones.filter(paciente__fecha_nacimiento__gte=date.today() - timedelta(days=365*10)).count(),
        '11 a 20 años': sesiones.filter(paciente__fecha_nacimiento__gte=date.today() - timedelta(days=365*20), paciente__fecha_nacimiento__lt=date.today() - timedelta(days=365*10)).count(),
        '21 a 30 años': sesiones.filter(paciente__fecha_nacimiento__gte=date.today() - timedelta(days=365*30), paciente__fecha_nacimiento__lt=date.today() - timedelta(days=365*20)).count(),
    }

    # Crear sección de Rango de Edades
    sheet['J6'] = "Rango de Edades"
    sheet['J6'].font = Font(bold=True, size=12)
    sheet['J6'].alignment = Alignment(horizontal='center')

    sheet['J7'] = "Rango"
    sheet['K7'] = "Cantidad"
    row_idx = 8
    for rango, count in rangos_edad.items():
        sheet[f'J{row_idx}'] = rango
        sheet[f'K{row_idx}'] = count
        row_idx += 1
    sheet[f'J{row_idx}'] = "Total"
    sheet[f'K{row_idx}'] = sum(rangos_edad.values())

    # Estilo para la tabla de edades
    for col in range(10, 12):
        for row in range(6, row_idx + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border_style

    # Crear gráfico de pastel para rango de edades
    chart_edades = PieChart()
    data_edades = Reference(sheet, min_col=11, min_row=8, max_row=row_idx - 1)
    categories_edades = Reference(sheet, min_col=10, min_row=8, max_row=row_idx - 1)
    chart_edades.add_data(data_edades, titles_from_data=False)
    chart_edades.set_categories(categories_edades)
    chart_edades.title = "Rango de edades de pacientes atendidos en Clínica Ompedis Ostuncalco"
    sheet.add_chart(chart_edades, "M15")

    # Calcular sesiones realizadas por semana
    sesiones_por_semana = sesiones.annotate(week=Week('fecha_sesion')).values('week').annotate(total=Count('id')).order_by('week')

    # Crear sección de Sesiones Realizadas por Semana
    sheet['J20'] = "Sesiones Realizadas por Semana"
    sheet['J20'].font = Font(bold=True, size=12)
    sheet['J20'].alignment = Alignment(horizontal='center')

    sheet['J21'] = "Semana"
    sheet['K21'] = "Sesiones Realizadas"
    row_idx = 22
    for sesion in sesiones_por_semana:
        sheet[f'J{row_idx}'] = f'Semana {int(sesion["week"])}'
        sheet[f'K{row_idx}'] = sesion['total']
        row_idx += 1
    sheet[f'J{row_idx}'] = "Total"
    sheet[f'K{row_idx}'] = sum(sesion['total'] for sesion in sesiones_por_semana)

    # Estilo para la tabla de sesiones por semana
    for col in range(10, 12):
        for row in range(20, row_idx + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border_style

    # Crear gráfico de barras para sesiones por semana
    chart_sesiones = BarChart()
    data_sesiones = Reference(sheet, min_col=11, min_row=22, max_row=row_idx - 1)
    categories_sesiones = Reference(sheet, min_col=10, min_row=22, max_row=row_idx - 1)
    chart_sesiones.add_data(data_sesiones, titles_from_data=False)
    chart_sesiones.set_categories(categories_sesiones)
    chart_sesiones.title = "Total de terapias dadas a pacientes atendidos por semana en Clínica Ompedis Ostuncalco"
    chart_sesiones.x_axis.title = "Semana"
    chart_sesiones.y_axis.title = "Total"
    sheet.add_chart(chart_sesiones, "M30")

    # Guardar el archivo en memoria
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename='reporte_sesiones_ompedis.xlsx')

@login_required
def historial_sesiones_view(request):
    sesiones = SesionTerapia.objects.all()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio:
        sesiones = sesiones.filter(fecha_sesion__gte=fecha_inicio)
    if fecha_fin:
        sesiones = sesiones.filter(fecha_sesion__lte=fecha_fin)

    sesiones = sesiones.order_by('fecha_sesion')

    context = {
        'sesiones': sesiones,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    return render(request, 'reportes/historial_sesiones.html', context)

class SesionTerapiaUpdateView(UpdateView):
    model = SesionTerapia
    form_class = SesionTerapiaForm
    template_name = 'reportes/editar_sesion.html'
    success_url = reverse_lazy('historial_sesiones')

    def form_valid(self, form):
        messages.success(self.request, 'La sesión de terapia se ha actualizado con éxito.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al actualizar la sesión de terapia. Por favor, verifique los datos ingresados.')
        return super().form_invalid(form)


class SesionTerapiaDeleteView(SuccessMessageMixin, DeleteView):
    model = SesionTerapia
    template_name = 'reportes/eliminar_sesion.html'
    success_url = reverse_lazy('historial_sesiones')
    success_message = 'La sesión de terapia se ha eliminado con éxito.'