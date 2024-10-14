from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import Paciente, Municipio
from .forms import PacienteForm
from django.http import JsonResponse
from django.db.models import Q
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.http import HttpResponse
import json
import openpyxl
from django import forms
from django.contrib import messages
from openpyxl.styles import Font, Alignment, Border, Side

# importaciones de decoradores
from django.utils.decorators import method_decorator
from usuarios.decorators import administrador_required, moderador_required, usuario_required, moderador_o_administrador_required
from django.contrib.auth.mixins import LoginRequiredMixin

@method_decorator(login_required, name='dispatch')
@method_decorator(moderador_o_administrador_required, name='dispatch')
class CrearPacienteView(CreateView):
    model = Paciente
    form_class = PacienteForm
    template_name = 'pacientes/crear_paciente.html'
    success_url = reverse_lazy('pacientes:lista_pacientes')

    def form_valid(self, form):
        messages.success(self.request, 'Se agregó el paciente con éxito.')
        return super().form_valid(form)

@method_decorator(login_required, name='dispatch')
@method_decorator(moderador_o_administrador_required, name='dispatch')
class EditarPacienteView(UpdateView):
    model = Paciente
    form_class = PacienteForm
    template_name = 'pacientes/editar_paciente.html'
    success_url = reverse_lazy('pacientes:lista_pacientes')

    def form_valid(self, form):
        messages.success(self.request, 'Se actualizó el paciente con éxito.')
        return super().form_valid(form)


def lista_pacientes_view(request):
    estado = request.GET.get('estado', 'activos')
    query = request.GET.get('q', '')
    if estado == 'inactivos':
        pacientes = Paciente.objects.filter(estado_activo=False)
    else:
        pacientes = Paciente.objects.filter(estado_activo=True)

    if query:
        pacientes = pacientes.filter(Q(nombre__icontains=query) | Q(apellido__icontains=query))
    pacientes= pacientes.order_by('nombre', 'apellido')
    context = {
        'pacientes': pacientes,
        'estado': estado,
        'query': query,
    }
    return render(request, 'pacientes/lista_pacientes.html', context)

@login_required
def cargar_municipios(request):
    departamento_id = request.GET.get('departamento')
    municipios = Municipio.objects.filter(departamento_id=departamento_id).order_by('nombre')
    html = render_to_string('pacientes/municipios_dropdown_list_options.html', {'municipios': municipios})
    return HttpResponse(html)


@method_decorator(login_required, name='dispatch')
class PacienteDetailView(DetailView):
    model = Paciente
    template_name = 'pacientes/detalle_paciente.html'
    context_object_name = 'paciente'




@moderador_o_administrador_required
def confirmar_cambio_estado(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk)
    context = {
        'paciente': paciente,
    }
    return render(request, 'pacientes/confirmar_cambio_estado.html', context)

@moderador_o_administrador_required
@require_POST
@csrf_exempt
def cambiar_estado_paciente_view(request):
    try:
        data = json.loads(request.body)
        paciente_id = data.get('id')
        nuevo_estado = data.get('estado') == 'activo'

        paciente = Paciente.objects.get(id=paciente_id)
        paciente.estado_activo = nuevo_estado
        paciente.save()

        # Filtrar los pacientes según el estado actual para actualizar la lista
        estado = request.GET.get('estado', 'activos')
        if estado == 'inactivos':
            pacientes = Paciente.objects.filter(estado_activo=False)
        else:
            pacientes = Paciente.objects.filter(estado_activo=True)

        query = request.GET.get('q', '')
        if query:
            pacientes = pacientes.filter(Q(nombre__icontains=query) | Q(apellido__icontains=query))

        context = {
            'pacientes': pacientes,
            'estado': estado,
            'query': query,
        }

        # Retornar la lista de pacientes actualizada
        return render(request, 'pacientes/lista_pacientes.html', context)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



@login_required
def exportar_pacientes_excel(request):
    estado = request.GET.get('estado', 'activos')
    
    # Filtrar pacientes según el estado activo o inactivo
    if estado == 'inactivos':
        pacientes = Paciente.objects.filter(estado_activo=False)
        estado_titulo = 'INACTIVOS'
    else:
        pacientes = Paciente.objects.filter(estado_activo=True)
        estado_titulo = 'ACTIVOS'

    pacientes = pacientes.order_by('nombre', 'apellido')

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    # Definir estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    align_center = Alignment(horizontal='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Añadir título
    ws.merge_cells('A1:O1')
    ws['A1'] = f"LISTA DE PACIENTES ({estado_titulo})"
    ws['A1'].font = title_font
    ws['A1'].alignment = align_center

    # Escribir los encabezados
    headers = [
        'Nombre', 'Apellido', 'ID Partida Nacimiento', 'Fecha Nacimiento', 'Género', 'Estado Activo', 
        'Departamento', 'Municipio', 'Domicilio', 'Diagnóstico Médico', 'Medicamentos', 
        'Responsable Nombre', 'Responsable Apellido', 'Responsable Parentesco', 'Responsable Teléfono'
    ]
    ws.append(headers)

    # Aplicar estilo a los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_style

    # Escribir los datos de los pacientes
    for row_num, paciente in enumerate(pacientes, start=3):
        ws.append([
            paciente.nombre,
            paciente.apellido,
            paciente.id_partida_nacimiento,
            paciente.fecha_nacimiento.strftime('%Y-%m-%d'),
            paciente.genero,
            'Activo' if paciente.estado_activo else 'Inactivo',
            paciente.departamento.nombre if paciente.departamento else '',
            paciente.municipio.nombre if paciente.municipio else '',
            paciente.domicilio,
            paciente.diagnostico_medico,
            paciente.medicamentos,
            paciente.responsable_nombre,
            paciente.responsable_apellido,
            paciente.responsable_parentesco,
            paciente.responsable_telefono
        ])
        # Aplicar estilo a las celdas de datos
        for col_num in range(1, 16):  # Desde la columna 1 hasta la 15
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border_style

    # Ajustar el ancho de las columnas para una mejor visualización
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col_letter].width = 20

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=pacientes_{estado}.xlsx'
    wb.save(response)
    return response
